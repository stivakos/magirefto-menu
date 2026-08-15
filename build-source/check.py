# -*- coding: utf-8 -*-
"""
Έλεγχος του DAILY_MENU.xlsx και του menu-today.txt πριν το ανέβασμα.

Πιάνει τις σιωπηλές αστοχίες — αυτές που ΔΕΝ ρίχνουν το build, αλλά κάνουν
ένα πιάτο να εξαφανιστεί από το site χωρίς να το καταλάβεις.

Τρέξε το από τον φάκελο build-source:      python check.py
Έξοδος: 0 = όλα καλά (ή μόνο προειδοποιήσεις), 1 = υπάρχει σφάλμα.
"""

import ast
import datetime
import os
import re
import sys
import unicodedata

import openpyxl

import dish_names          # ίδια αναγνώριση ονομάτων με το build.py
from dish_names import norm, sound, stems
import menu_date           # ίδιος parser ημερομηνίας με το build.py

HERE = os.path.dirname(os.path.abspath(__file__))
BUILD_PY = os.path.join(HERE, "build.py")
XLSX = os.path.join(HERE, "..", "DAILY_MENU.xlsx")
MENU_TXT = os.path.join(HERE, "..", "menu-today.txt")


# --- ρυθμίσεις: διαβάζονται από το build.py ώστε να μην ξεφύγουν ποτέ ------
WANTED = ("CATEGORIES", "HIDE_PRICE", "GALLERY", "GALLERY_DIR", "SIDE_YES")
# «όχι» δεν το θεωρούμε λάθος — είναι ο φυσικός τρόπος να γράψεις «δεν παίρνει».
SIDE_NO = {"όχι", "οχι", "ο", "-", "—", "n/a", "no", "0"}


def config_from_build():
    """Παίρνει τις σταθερές από το build.py χωρίς να το εκτελέσει."""
    tree = ast.parse(open(BUILD_PY, encoding="utf-8").read())
    found = {}
    for node in tree.body:
        if isinstance(node, ast.Assign) and len(node.targets) == 1:
            t = node.targets[0]
            if isinstance(t, ast.Name) and t.id in WANTED:
                try:
                    found[t.id] = ast.literal_eval(node.value)
                except ValueError:
                    pass
    missing = set(WANTED) - set(found)
    if missing:
        sys.exit(f"!! Δεν βρέθηκαν {missing} στο build.py — άλλαξε η δομή του;")
    return (found["CATEGORIES"], set(found["HIDE_PRICE"]),
            found["GALLERY"], found["GALLERY_DIR"], set(found["SIDE_YES"]))


CATEGORIES, HIDE_PRICE, GALLERY, GALLERY_DIR, SIDE_YES = config_from_build()
with_side = []          # πιάτα σημαδεμένα «Με συνοδευτικό;»

errors, warnings = [], []


def err(msg):
    errors.append(msg)


def warn(msg):
    warnings.append(msg)


# --- εντοπισμός ορθογραφικών παραλλαγών του ίδιου πιάτου -------------------
# «Ρεβίθια» vs «Ρεβύθια», «Γλώσσα» vs «Γλώσσες»: ίδιο πιάτο, δύο γραφές, δύο
# Α/Α — και το site τα δείχνει σαν διαφορετικά. Ο έλεγχος ισοπεδώνει τους
# ήχους που στα ελληνικά γράφονται με πολλούς τρόπους (ι/η/υ/ει/οι, ο/ω,
# ε/αι), διπλά σύμφωνα και καταλήξεις, και μετά συγκρίνει.
def dist(a, b, cap=3):
    """Απόσταση Levenshtein, με πρόωρη έξοδο πάνω από το cap."""
    if abs(len(a) - len(b)) > cap:
        return cap + 1
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1,
                           prev[j - 1] + (ca != cb)))
        if min(cur) > cap:
            return cap + 1
        prev = cur
    return prev[-1]


def similar_pairs(items):
    """items: [(Α/Α, όνομα)] -> ζεύγη που μοιάζουν ύποπτα."""
    out = []
    keys = [(n, name, sound(name), stems(name)) for n, name in items]
    for i in range(len(keys)):
        n1, name1, s1, st1 = keys[i]
        for j in range(i + 1, len(keys)):
            n2, name2, s2, st2 = keys[j]
            if st1 and st1 == st2 and s1 != s2:
                out.append((n1, name1, n2, name2, "ίδιες ρίζες λέξεων"))
                continue
            if not s1 or not s2 or abs(len(s1) - len(s2)) > 3:
                continue
            # «Νερό 1lt» vs «Νερό 1,5lt»: παραλλαγή μεγέθους, όχι ορθογραφίας
            if (re.findall(r"\d+", name1) != re.findall(r"\d+", name2)
                    and re.sub(r"[\d,.]+", "", s1) == re.sub(r"[\d,.]+", "", s2)):
                continue
            if s1 == s2:
                out.append((n1, name1, n2, name2, "ίδια προφορά"))
            elif len(s1) >= 5 and dist(s1, s2) <= 2:
                out.append((n1, name1, n2, name2, "σχεδόν ίδια"))
    return out


# --- 1. το ίδιο το αρχείο --------------------------------------------------
if not os.path.exists(XLSX):
    sys.exit(f"!! Δεν βρέθηκε το {XLSX}")

wb = openpyxl.load_workbook(XLSX, data_only=True)

tabs_needed = [tab for _, _, tab in CATEGORIES]
for tab in tabs_needed:
    if tab not in wb.sheetnames:
        err(f"ΛΕΙΠΕΙ ΤΟ TAB «{tab}» — το build θα σκάσει. "
            f"Υπάρχοντα tabs: {', '.join(wb.sheetnames)}")

extra = [s for s in wb.sheetnames if s not in tabs_needed]
if extra:
    warn(f"Tabs που αγνοούνται εντελώς από το site: {', '.join(extra)}")

if errors:
    print("\n".join("✗ " + e for e in errors))
    sys.exit(1)


# --- 2. περιεχόμενο κάθε tab ----------------------------------------------
catalog = {}   # slug -> {Α/Α: (όνομα, τιμή)}
summary = []

for label, slug, tab in CATEGORIES:
    ws = wb[tab]
    rows, names_seen = {}, {}
    valid = 0

    for r in range(2, ws.max_row + 1):
        aa = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        price = ws.cell(r, 3).value

        blank = aa is None and not name and price in (None, "")
        if blank:
            continue

        # Α/Α
        if aa is None:
            if name:
                err(f"[{tab}] γραμμή {r}: το πιάτο «{str(name).strip()}» "
                    f"ΔΕΝ ΕΧΕΙ Α/Α — δεν θα εμφανιστεί ποτέ.")
            continue
        try:
            n = int(aa)
        except (TypeError, ValueError):
            err(f"[{tab}] γραμμή {r}: Α/Α «{aa}» δεν είναι ακέραιος — "
                f"η γραμμή προσπερνιέται σιωπηλά.")
            continue

        # όνομα
        if not name or not str(name).strip():
            err(f"[{tab}] γραμμή {r}: Α/Α {n} χωρίς ονομασία — προσπερνιέται.")
            continue
        name = str(name).strip()

        # τιμή — η κλασική παγίδα: κείμενο «8,50» ρίχνει ΟΛΗ τη γραμμή
        val = None
        if price not in (None, ""):
            if isinstance(price, str):
                if "," in price:
                    err(f"[{tab}] γραμμή {r}: «{name}» — τιμή «{price}» είναι "
                        f"ΚΕΙΜΕΝΟ με κόμμα. Το πιάτο ΘΑ ΕΞΑΦΑΝΙΣΤΕΙ. "
                        f"Γράψε την ως αριθμό ({price.replace(',', '.')}).")
                    continue
                try:
                    val = float(price)
                    warn(f"[{tab}] γραμμή {r}: «{name}» — η τιμή είναι κείμενο "
                         f"«{price}». Δουλεύει, αλλά κάν' την αριθμό.")
                except ValueError:
                    err(f"[{tab}] γραμμή {r}: «{name}» — τιμή «{price}» δεν "
                        f"διαβάζεται ως αριθμός. Το πιάτο θα εξαφανιστεί.")
                    continue
            else:
                try:
                    val = float(price)
                except (TypeError, ValueError):
                    err(f"[{tab}] γραμμή {r}: «{name}» — τιμή «{price}» άκυρη. "
                        f"Το πιάτο θα εξαφανιστεί.")
                    continue
            if val is not None and val <= 0:
                warn(f"[{tab}] γραμμή {r}: «{name}» με τιμή {val}.")
        elif slug not in HIDE_PRICE:
            warn(f"[{tab}] γραμμή {r}: «{name}» χωρίς τιμή — θα εμφανιστεί "
                 f"στο site χωρίς τιμή.")

        # στήλη E «Με συνοδευτικό;» — άγνωστη τιμή = σιωπηλά αγνοείται από
        # το build, οπότε ο ιδιοκτήτης νομίζει ότι το σήμανε και δεν έγινε.
        side_raw = ws.cell(r, 5).value if ws.max_column >= 5 else None
        side_txt = str(side_raw or "").strip().lower()
        if side_txt and side_txt not in SIDE_YES and side_txt not in SIDE_NO:
            warn(f"[{tab}] γραμμή {r}: «{name}» — στήλη «Με συνοδευτικό;» έχει "
                 f"«{side_raw}», που δεν αναγνωρίζεται. Γράψε «Ν» ή άφησέ το κενό.")
        elif side_txt in SIDE_YES:
            with_side.append(name)

        # διπλά
        if n in rows:
            err(f"[{tab}] γραμμή {r}: ΔΙΠΛΟ Α/Α {n} («{name}» και "
                f"«{rows[n][0]}»). Το ένα από τα δύο χάνεται.")
        rows[n] = (name, val)

        key = norm(name)
        if key in names_seen:
            warn(f"[{tab}]: ίδια ονομασία σε Α/Α {names_seen[key]} και {n} "
                 f"(«{name}»).")
        else:
            names_seen[key] = n
        valid += 1

    for n1, name1, n2, name2, why in similar_pairs(
            [(n, v[0]) for n, v in rows.items()]):
        warn(f"[{tab}]: «{name1}» (Α/Α {n1}) και «{name2}» (Α/Α {n2}) — "
             f"{why}. Ίδιο πιάτο γραμμένο αλλιώς; Τότε κράτα το ένα Α/Α.")

    catalog[slug] = rows
    nxt = (max(rows) + 1) if rows else 1
    gaps = [i for i in range(1, max(rows) + 1) if i not in rows] if rows else []
    summary.append((label, valid, nxt, gaps))


# --- 3. η επιλογή της ημέρας ----------------------------------------------
slug_by_norm = {dish_names.category_key(lbl): slug
                for lbl, slug, _ in CATEGORIES}
label_by_slug = {slug: lbl for lbl, slug, _ in CATEGORIES}
date_found = False
closed = ""
picked = {}

if not os.path.exists(MENU_TXT):
    err(f"Δεν βρέθηκε το {MENU_TXT}")
else:
    for ln, raw in enumerate(open(MENU_TXT, encoding="utf-8"), 1):
        line = raw.strip()
        if not line or line.startswith("#") or ":" not in line:
            continue
        key, val = line.split(":", 1)
        kn = dish_names.category_key(key)
        if kn in ("ημερομηνια", "date"):
            date_found = bool(val.strip())
            if not date_found and not closed:
                err(f"menu-today.txt γραμμή {ln}: κενή ΗΜΕΡΟΜΗΝΙΑ.")
            menu_day = menu_date.parse(val)
            if date_found and not menu_day:
                warn(f"menu-today.txt γραμμή {ln}: η ημερομηνία «{val.strip()}» "
                     f"δεν διαβάζεται ως ημέρα/μήνας. Η σελίδα δεν θα μπορεί να "
                     f"καταλάβει αν το μενού είναι σημερινό.")
            elif menu_day:
                today = datetime.date.today()
                if menu_day < today:
                    days = (today - menu_day).days
                    warn(f"Η ΗΜΕΡΟΜΗΝΙΑ είναι {days} μέρα(ες) πίσω "
                         f"({menu_day.strftime('%d/%m/%Y')}). Η σελίδα θα βγάλει "
                         f"«δεν ενημερώθηκε για σήμερα» και ΔΕΝ θα δέχεται "
                         f"παραγγελίες.")
                elif menu_day > today:
                    warn(f"Η ΗΜΕΡΟΜΗΝΙΑ είναι μελλοντική "
                         f"({menu_day.strftime('%d/%m/%Y')}) — εντάξει αν "
                         f"ετοιμάζεις το αυριανό μενού.")
                bad = menu_date.weekday_mismatch(val)
                if bad:
                    warn(f"Έγραψες «{bad[0]}» αλλά η "
                         f"{menu_day.strftime('%d/%m/%Y')} είναι {bad[1]}.")
            continue
        if kn in ("κλειστα", "closed"):
            closed = val.strip()
            if not closed:
                err(f"menu-today.txt γραμμή {ln}: η γραμμή ΚΛΕΙΣΤΑ είναι κενή. "
                    f"Γράψε πότε ανοίγεις (π.χ. «ΚΛΕΙΣΤΑ: Τρίτη 25/8») ή "
                    f"σβήσε τη γραμμή για να ανοίξει το μαγαζί.")
            continue
        if kn not in slug_by_norm:
            err(f"menu-today.txt γραμμή {ln}: άγνωστη κατηγορία «{key.strip()}» "
                f"— αγνοείται. Επιτρεπτές: "
                f"{', '.join(l for l, _, _ in CATEGORIES)}")
            continue

        slug = slug_by_norm[kn]
        rows = catalog.get(slug, {})
        # δέχεται αριθμούς, ονόματα («μπιφτέκι, μουσακάς») ή μείγμα
        nums, name_errs = dish_names.parse_selection(val, rows)
        for m in name_errs:
            err(f"menu-today.txt γραμμή {ln} ({label_by_slug[slug]}): {m}")
        picked[slug] = nums

        seen = set()
        for n in nums:
            if n in seen:
                warn(f"menu-today.txt γραμμή {ln} ({label_by_slug[slug]}): "
                     f"ο αριθμός {n} γραμμένος δύο φορές.")
            seen.add(n)
            if n not in rows:
                err(f"menu-today.txt γραμμή {ln} ({label_by_slug[slug]}): "
                    f"Α/Α {n} ΔΕΝ ΥΠΑΡΧΕΙ στο xlsx — αγνοείται σιωπηλά.")

    if not date_found and not closed:
        err("menu-today.txt: λείπει εντελώς η γραμμή ΗΜΕΡΟΜΗΝΙΑ.")

    # όσο το μαγαζί είναι κλειστό, το μενού δεν εμφανίζεται — μην γκρινιάζεις
    if not closed:
        for _, slug, _ in CATEGORIES:
            if not picked.get(slug):
                warn(f"Η κατηγορία «{label_by_slug[slug]}» δεν έχει επιλογές — "
                     f"δεν θα εμφανιστεί σήμερα.")


    # Σημαδεμένα πιάτα χωρίς συνοδευτικά της ημέρας = η επιλογή δεν εμφανίζεται
    if with_side and not picked.get("synodeytika") and not closed:
        warn(f"{len(with_side)} πιάτα δέχονται συνοδευτικό, αλλά σήμερα δεν "
             f"έχεις επιλέξει κανένα Συνοδευτικό — η επιλογή δεν θα φανεί.")


# --- 5. οι φωτογραφίες της γκαλερί -----------------------------------------
# Λείπον αρχείο δεν χαλάει το build — βγάζει σπασμένη εικόνα στο live site, που
# κανείς δεν βλέπει μέχρι να το πει πελάτης.
GALLERY_PATH = os.path.join(HERE, "..", *GALLERY_DIR.split("/"))
listed = [f for f, _, _ in GALLERY]
gallery_kb = 0

if not os.path.isdir(GALLERY_PATH):
    err(f"Δεν βρέθηκε ο φάκελος {GALLERY_DIR}/ — τρέξε το prep-photos.sh.")
else:
    for f in listed:
        p = os.path.join(GALLERY_PATH, f)
        if not os.path.isfile(p):
            err(f"Γκαλερί: λείπει η φωτογραφία {GALLERY_DIR}/{f} — "
                f"τρέξε το prep-photos.sh ή διόρθωσε το GALLERY στο build.py.")
        else:
            gallery_kb += os.path.getsize(p) / 1024
    for f in sorted(os.listdir(GALLERY_PATH)):
        if not f.startswith(".") and f not in listed:
            warn(f"Γκαλερί: το {GALLERY_DIR}/{f} δεν χρησιμοποιείται πουθενά.")


# --- 6. αναφορά ------------------------------------------------------------
print("\n── ΒΑΣΗ ΠΙΑΤΩΝ (DAILY_MENU.xlsx) " + "─" * 28)
for label, valid, nxt, gaps in summary:
    g = f"  κενά Α/Α: {', '.join(map(str, gaps))}" if gaps else ""
    print(f"  {label:20s} {valid:3d} πιάτα   επόμενο Α/Α: {nxt}{g}")

if closed:
    print("\n── ΚΛΕΙΣΤΟ ΜΑΓΑΖΙ " + "─" * 42)
    print(f"  Το site δείχνει «Ανοίγουμε {closed}» — χωρίς μενού και χωρίς")
    print(f"  παραγγελίες. Σβήσε τη γραμμή ΚΛΕΙΣΤΑ για να επανέλθει το μενού.")

if picked and not closed:
    print("\n── ΜΕΝΟΥ ΗΜΕΡΑΣ " + "─" * 44)
    for _, slug, _ in CATEGORIES:
        nums = picked.get(slug, [])
        if not nums:
            continue
        rows = catalog.get(slug, {})
        shown = [rows[n][0] for n in nums if n in rows]
        print(f"  {label_by_slug[slug]:20s} {len(shown)}: "
              f"{', '.join(shown) if shown else '—'}")

if with_side:
    print("\n── ΜΕ ΣΥΝΟΔΕΥΤΙΚΟ " + "─" * 42)
    print(f"  {len(with_side)}: {', '.join(sorted(with_side))}")

print("\n── ΤΟ ΜΑΓΑΖΙ ΜΑΣ " + "─" * 43)
print(f"  {len(listed)} φωτογραφίες, {gallery_kb / 1024:.1f} MB συνολικά.")

if warnings:
    print("\n── ΠΡΟΕΙΔΟΠΟΙΗΣΕΙΣ " + "─" * 41)
    for w in warnings:
        print("  ⚠  " + w)

if errors:
    print("\n── ΣΦΑΛΜΑΤΑ " + "─" * 48)
    for e in errors:
        print("  ✗  " + e)
    print(f"\n{len(errors)} σφάλμα(τα). Διόρθωσέ τα πριν το ανέβασμα.\n")
    sys.exit(1)

print(f"\n✓ Όλα εντάξει"
      f"{f' ({len(warnings)} προειδοποιήσεις)' if warnings else ''}.\n")
