#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the Merci Μαγειρευτό online menu (index.html) + DAILY_MENU.xlsx.

Single source of truth = MENU below. Categories are fixed; dishes edited on
request. Prices come from Μαγειρευτό_Μενού.xlsx (owner's rule). Price = number
or None (side dishes shown without price).
"""
import re, html, os

HERE = os.path.dirname(os.path.abspath(__file__))
ONIRO = "/Users/stavros/oniropetra-menu/index.html"
OUT = os.path.join(HERE, "..", "index.html")
XLSX_OUT = "/Users/stavros/Documents/ΜΑΓΕΙΡΕΥΤΟ/DAILY_MENU.xlsx"

FONT_FACES = "\n".join(re.findall(r'@font-face\s*\{.*?\}',
                                  open(ONIRO, encoding="utf-8").read(), re.S))

MENU_DATE = "Δευτέρα 20/7/26"   # η ημερομηνία που αφορά το μενού

# category = dict(slug, label, note, items[]) ; item = dict(name, price, desc, portion)
MENU = [
    {"slug": "mageirefta", "label": "Μαγειρευτά", "items": [
        {"name": "Κοτόσουπα",                        "price": 7.5},
        {"name": "Γεμιστά",                          "price": 6.5},
        {"name": "Παστίτσιο",                        "price": 8.5},
        {"name": "Μουσακάς",                         "price": 8.5},
        {"name": "Γίγαντες Καστοριάς",               "price": 6.7},
        {"name": "Μοσχάρι γιουβέτσι",                "price": 9.5},
        {"name": "Μοσχάρι με αρακά",                 "price": 9.5},
        {"name": "Χταπόδι κριθαρότο",                "price": 10.5},
        {"name": "Σουφλέ ζυμαρικών",                 "price": 7.5},
        {"name": "Κεφτεδάκια τηγανητά & κοκκινιστά", "price": 7.5},
    ]},
    {"slug": "tis-oras", "label": "Της ώρας", "items": [
        {"name": "Σνίτσελ κοτόπουλο",                "price": 8.5},
        {"name": "Φιλέτο κοτόπουλο",                 "price": 8.0},
        {"name": "Καλαμαράκια",                      "price": 9.0},
        {"name": "Μπακαλιάρος τηγανητός",            "price": 10.0},
        {"name": "Σολομός φούρνου",                  "price": 12.0},
        {"name": "Φιλετίνια κοτόπουλο a la creme",   "price": 7.5},
    ]},
    {"slug": "synodeytika", "label": "Συνοδευτικά",
     "note": "…και σε μερίδα για μεγαλύτερη απόλαυση!", "items": [
        {"name": "Πατάτες τηγανητές",               "price": None},
        {"name": "Πατάτες ψητές",                   "price": None},
        {"name": "Πουρές πατάτας",                  "price": None},
        {"name": "Ρύζι",                            "price": None},
    ]},
    {"slug": "salates", "label": "Σαλάτες", "items": [
        {"name": "Αγγουροντομάτα",                  "price": 3.5},
        {"name": "Χωριάτικη",                       "price": 5.0},
        {"name": "Λάχανο - Καρότο",                 "price": 3.0},
        {"name": "Μαρούλι",                         "price": 3.5},
        {"name": "Παντζάρι",                        "price": 3.5},
        {"name": "Κουνουπίδι",                      "price": 3.5},
        {"name": "Φέτα (Π.Ο.Π.)",                   "price": 4.0},
        {"name": "Φρουτοσαλάτα εποχής",             "price": 4.5},
    ]},
    {"slug": "glyka", "label": "Γλυκά", "items": [
        {"name": "Πορτοκαλόπιτα",                   "price": 4.0},
        {"name": "Σοκολατόπιτα",                    "price": 5.0},
    ]},
    {"slug": "anapsyktika", "label": "Αναψυκτικά / Ποτά", "items": [
        {"name": "Νερό 330ml",                      "price": 0.5},
        {"name": "Νερό 1,5lt",                      "price": 1.5},
        {"name": "Coca-Cola / Zero / Light 330ml",  "price": 1.8},
    ]},
]

def esc(s): return html.escape(str(s), quote=True)
def fmt_price(v):
    return None if v is None else f"{float(v):.2f}".replace(".", ",") + " €"

def item_html(it):
    portion = f' <span class="portion">{esc(it["portion"])}</span>' if it.get("portion") else ""
    p = fmt_price(it.get("price"))
    if p:
        line = (f'<div class="item-line"><span class="gr">{esc(it["name"])}{portion}</span>'
                f'<span class="dots"></span><span class="price">{esc(p)}</span></div>')
    else:
        line = f'<div class="item-line"><span class="gr">{esc(it["name"])}{portion}</span></div>'
    if it.get("desc"):
        line += f'\n        <p class="desc" lang="el">{esc(it["desc"])}</p>'
    return f'      <li class="item">{line}</li>'

nav = "\n".join(f'    <a class="chip" href="#{c["slug"]}">{esc(c["label"])}</a>' for c in MENU)

secs = []
for c in MENU:
    if c["items"]:
        body = '<ul class="items">\n' + "\n".join(item_html(it) for it in c["items"]) + '\n    </ul>'
    else:
        body = '<p class="empty-note">— σύντομα —</p>'
    if c.get("note"):
        body += f'\n    <p class="sec-note" lang="el">{esc(c["note"])}</p>'
    secs.append(f'''  <section id="{c["slug"]}" aria-labelledby="h-{c["slug"]}">
    <div class="sec-head">
      <h2 id="h-{c["slug"]}" lang="el">{esc(c["label"])}</h2>
    </div>
    {body}
  </section>''')
sections_html = "\n\n".join(secs)

CSS = """
  :root{
    --paper:#0E1D34; --raised:#152740; --ink:#F1E7D5; --muted:#B9C4D4; --faint:#7F90A6;
    --sea:#E0885A; --sand:#D9BE73; --sand-deep:#CBA85C; --leader:#33445E;
    --hairline:#26374F; --chip-bg:#17293F; --mist-2:#0B1728; --pot:#E0885A;
    --display:"GFS Didot","Palatino Linotype",Palatino,Georgia,serif;
    --body:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  }
  *{box-sizing:border-box;} html{-webkit-text-size-adjust:100%;}
  @media (prefers-reduced-motion:no-preference){html{scroll-behavior:smooth;}}
  body{margin:0;background:var(--paper);color:var(--ink);font-family:var(--body);font-size:1rem;line-height:1.55;}

  .cove{position:relative;overflow:hidden;background:linear-gradient(160deg,#122540,var(--paper));text-align:center;padding:3.25rem 1.25rem 2.2rem;border-bottom:1px solid var(--hairline);}
  .pot{position:absolute;top:1.6rem;left:6%;width:clamp(78px,17vw,128px);height:auto;}
  .pan{position:absolute;top:2.2rem;right:6%;width:clamp(74px,16vw,120px);height:auto;}
  .steam{fill:none;stroke:#EAD9BE;stroke-width:4;stroke-linecap:round;opacity:.0;transform-origin:center;}
  @media (prefers-reduced-motion:no-preference){
    .steam{animation:steam 3.4s ease-in-out infinite;}
    .steam.s2{animation-delay:.7s;} .steam.s3{animation-delay:1.4s;}
  }
  @media (prefers-reduced-motion:reduce){ .steam{opacity:.5;} }
  @keyframes steam{0%{opacity:0;transform:translateY(6px) scaleY(.85);}
    35%{opacity:.75;} 70%{opacity:.35;} 100%{opacity:0;transform:translateY(-6px) scaleY(1.1);}}

  .brand{position:relative;font-family:var(--display);font-weight:400;font-size:clamp(2.6rem,10vw,4.2rem);line-height:1.05;letter-spacing:.01em;margin:0;text-wrap:balance;}
  .brand .merci{display:block;font-size:.4em;letter-spacing:.5em;text-indent:.5em;text-transform:uppercase;color:var(--sea);margin:0 0 .3rem;}
  .brand-sub{position:relative;margin:.9rem 0 0;font-size:.74rem;font-weight:600;letter-spacing:.3em;text-indent:.3em;text-transform:uppercase;color:var(--sand);}
  .menu-date{position:relative;margin:.55rem 0 0;font-family:var(--display);font-size:clamp(1.15rem,4.5vw,1.5rem);color:var(--ink);}

  .rail{position:sticky;top:0;z-index:10;background:color-mix(in srgb,var(--paper) 90%,transparent);-webkit-backdrop-filter:blur(10px);backdrop-filter:blur(10px);border-bottom:1px solid var(--hairline);}
  .rail-inner{display:flex;gap:.5rem;overflow-x:auto;padding:.7rem 1.1rem;max-width:46rem;margin:0 auto;scrollbar-width:none;justify-content:flex-start;}
  .rail-inner::-webkit-scrollbar{display:none;}
  .chip{flex:0 0 auto;display:inline-flex;align-items:center;justify-content:center;min-height:44px;padding:.4rem 1.1rem;border-radius:18px;border:1px solid var(--hairline);background:var(--chip-bg);color:var(--muted);font-size:.94rem;font-weight:600;line-height:1.2;text-align:center;text-decoration:none;white-space:nowrap;cursor:pointer;transition:background-color .2s,color .2s,border-color .2s;}
  .chip:hover{border-color:var(--sea);color:var(--ink);}
  .chip.is-active{background:var(--sea);border-color:var(--sea);color:#10203A;}
  .chip:focus-visible,a:focus-visible{outline:2px solid var(--sand);outline-offset:2px;}

  main{max-width:44rem;margin:0 auto;padding:.75rem 1.25rem 2rem;}
  section{scroll-margin-top:4.6rem;padding-top:2.6rem;}
  .sec-head{display:flex;align-items:baseline;gap:.75rem;border-bottom:2px solid var(--sand-deep);padding-bottom:.55rem;}
  .sec-head h2{font-family:var(--display);font-weight:400;font-size:clamp(1.8rem,6vw,2.3rem);margin:0;letter-spacing:.01em;color:var(--ink);}
  .items{list-style:none;margin:0;padding:0;}
  .item{padding:.72rem 0;border-bottom:1px solid var(--hairline);}
  .item:last-child{border-bottom:none;}
  .item-line{display:flex;align-items:baseline;gap:.55rem;}
  .gr{font-weight:600;}
  .portion{font-weight:400;font-size:.8rem;color:var(--faint);}
  .dots{flex:1 1 1.5rem;min-width:1.5rem;border-bottom:2px dotted var(--leader);transform:translateY(-.28em);}
  .price{font-variant-numeric:tabular-nums;font-weight:650;white-space:nowrap;color:var(--sand);}
  .desc{margin:.3rem 0 0;font-size:.86rem;color:var(--muted);max-width:34rem;}
  .sec-note{margin:1rem 0 0;font-family:var(--display);font-size:1.05rem;font-style:italic;color:var(--sea);text-align:center;}
  .empty-note{margin:1.4rem 0 .4rem;color:var(--faint);font-style:italic;text-align:center;}

  footer{border-top:1px solid var(--hairline);background:var(--mist-2);text-align:center;padding:2.2rem 1.5rem 2.6rem;}
  .foot-brand{font-family:var(--display);font-size:1.4rem;margin:0 0 .2rem;}
  .foot-place{font-size:.72rem;font-weight:700;letter-spacing:.26em;text-indent:.26em;text-transform:uppercase;color:var(--sea);margin:0 0 1.2rem;}
  .foot-hours{font-size:.9rem;color:var(--ink);margin:0 0 1rem;}
  .legal{max-width:34rem;margin:0 auto;font-size:.76rem;line-height:1.6;color:var(--muted);}
  .legal p{margin:.35rem 0;}
"""

POT_SVG = '''<svg class="pot" viewBox="0 0 130 130" aria-hidden="true">
    <path class="steam s1" d="M50 44 C44 36 56 32 50 24 C44 16 56 12 52 6"/>
    <path class="steam s2" d="M67 44 C61 36 73 32 67 24 C61 16 73 12 69 6"/>
    <path class="steam s3" d="M84 44 C78 36 90 32 84 24 C78 16 90 12 86 6"/>
    <rect x="30" y="60" width="70" height="46" rx="9" fill="#C9975B"/>
    <rect x="30" y="60" width="70" height="14" rx="7" fill="#E0A96D"/>
    <rect x="22" y="52" width="86" height="12" rx="6" fill="#E0885A"/>
    <rect x="58" y="45" width="14" height="9" rx="4" fill="#E0885A"/>
    <rect x="14" y="72" width="12" height="20" rx="6" fill="#B07C3F"/>
    <rect x="104" y="72" width="12" height="20" rx="6" fill="#B07C3F"/>
  </svg>'''

PAN_SVG = '''<svg class="pan" viewBox="0 0 140 120" aria-hidden="true">
    <path class="steam s1" d="M52 40 C46 32 58 28 52 20 C46 12 58 8 54 2"/>
    <path class="steam s2" d="M70 40 C64 32 76 28 70 20 C64 12 76 8 72 2"/>
    <ellipse cx="60" cy="78" rx="46" ry="30" fill="#B07C3F"/>
    <ellipse cx="60" cy="73" rx="46" ry="30" fill="#C9975B"/>
    <ellipse cx="60" cy="71" rx="37" ry="22" fill="#2C2013"/>
    <ellipse cx="52" cy="65" rx="9" ry="5" fill="#E0885A" opacity=".7"/>
    <rect x="100" y="66" width="40" height="11" rx="5.5" fill="#7A5326" transform="rotate(-16 100 66)"/>
  </svg>'''

HTML = f'''<!doctype html>
<html lang="el">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Merci Μαγειρευτό · Μενού — Λάρισα</title>
<meta name="description" content="Μενού — Merci Μαγειρευτό, σπιτικό φαγητό, Λάρισα. Take away & delivery.">
</head>
<body>
<style>
{FONT_FACES}
{CSS}
</style>

<header class="cove">
  {POT_SVG}
  {PAN_SVG}
  <h1 class="brand" lang="el"><span class="merci">Merci</span>Μαγειρευτό</h1>
  <p class="brand-sub">Σπιτικό φαγητό</p>
  <p class="menu-date" lang="el">{esc(MENU_DATE)}</p>
</header>

<nav class="rail" aria-label="Κατηγορίες μενού">
  <div class="rail-inner">
{nav}
  </div>
</nav>

<main>
{sections_html}
</main>

<footer>
  <p class="foot-brand" lang="el">Merci Μαγειρευτό</p>
  <p class="foot-place">Λάρισα · Take away &amp; Delivery</p>
  <p class="foot-hours" lang="el">Δευτέρα – Σάββατο, 12:00 – 15:30</p>
  <div class="legal">
    <p lang="el">Οι τιμές περιλαμβάνουν όλους τους νόμιμους φόρους. Ο καταναλωτής δεν έχει την υποχρέωση να πληρώσει εάν δε λάβει το νόμιμο παραστατικό στοιχείο (απόδειξη-τιμολόγιο).</p>
  </div>
</footer>

<script>
  (function () {{
    var chips = Array.prototype.slice.call(document.querySelectorAll(".chip"));
    var byId = {{}};
    chips.forEach(function (chip) {{ byId[chip.getAttribute("href").slice(1)] = chip; }});
    var current = null;
    function activate(id) {{
      if (current === id) return;
      current = id;
      chips.forEach(function (chip) {{ chip.classList.remove("is-active"); }});
      var chip = byId[id];
      if (chip) {{ chip.classList.add("is-active"); chip.scrollIntoView({{ block: "nearest", inline: "center", behavior: "smooth" }}); }}
    }}
    if ("IntersectionObserver" in window) {{
      var visible = {{}};
      var observer = new IntersectionObserver(function (entries) {{
        entries.forEach(function (entry) {{ visible[entry.target.id] = entry.isIntersecting; }});
        var secs = document.querySelectorAll("main section");
        for (var i = 0; i < secs.length; i++) {{ if (visible[secs[i].id]) {{ activate(secs[i].id); break; }} }}
      }}, {{ rootMargin: "-20% 0px -60% 0px" }});
      document.querySelectorAll("main section").forEach(function (sec) {{ observer.observe(sec); }});
    }}
  }})();
</script>
</body>
</html>
'''

with open(OUT, "w", encoding="utf-8") as f:
    f.write(HTML)
print(f"Wrote {OUT}  ({len(MENU)} categories, {sum(len(c['items']) for c in MENU)} dishes)")

# ---------------------------------------------------------------------------
# DAILY_MENU.xlsx  — one tab per category: Α/Α | Ονομασία πιάτου | Τιμή
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    hdr_fill = PatternFill("solid", fgColor="1F3A5F"); hdr_font = Font(bold=True, color="F1E7D5")
    thin = Side(style="thin", color="D9D2C4"); border = Border(thin, thin, thin, thin)
    center = Alignment(horizontal="center", vertical="center")
    def safe(t): return re.sub(r'[\\/?*\[\]:]', "-", t)[:31]
    for c in MENU:
        ws = wb.create_sheet(safe(c["label"]))
        ws.append(["Α/Α", "Ονομασία πιάτου", "Τιμή (€)"])
        for i in range(1, 4):
            cell = ws.cell(1, i); cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border
        for idx, it in enumerate(c["items"], 1):
            ws.append([idx, it["name"], it.get("price")])
            ws.cell(ws.max_row, 1).alignment = center
            pc = ws.cell(ws.max_row, 3)
            if it.get("price") is not None: pc.number_format = '0.00" €"'
            pc.alignment = center
            for col in range(1, 4): ws.cell(ws.max_row, col).border = border
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 12
        ws.freeze_panes = "A2"
    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}  ({len(MENU)} tabs)")
except Exception as e:
    print("xlsx skipped:", e)
